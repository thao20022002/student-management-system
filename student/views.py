from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from .models import Parent, Student, Class, Subject, Grade, Attendance, Teacher, AdmissionCandidate, Schedule
import os
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q, Avg, Count
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from .decorators import admin_required, teacher_required, admin_or_teacher_required, student_required
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator

from student.services.grade_averages import round_average_score


@admin_required
def add_student(request):
    if request.method == "POST":
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        student_id = request.POST.get('student_id')
        gender = request.POST.get('gender')
        date_of_birth = request.POST.get('date_of_birth')
        student_class = request.POST.get('student_class')
        religion = request.POST.get('religion')
        joining_date = request.POST.get('joining_date')
        mobile_number = request.POST.get('mobile_number')
        admission_number = request.POST.get('admission_number')
        section = request.POST.get('section')
        student_image = request.FILES.get('student_image')

        father_name = request.POST.get('father_name')
        father_occupation = request.POST.get('father_occupation')
        father_mobile = request.POST.get('father_mobile')
        father_email = request.POST.get('father_email')
        mother_name = request.POST.get('mother_name')
        mother_occupation = request.POST.get('mother_occupation')
        mother_mobile = request.POST.get('mother_mobile')
        mother_email = request.POST.get('mother_email')
        present_address = request.POST.get('present_address')
        permanent_address = request.POST.get('permanent_address')

        parent = Parent.objects.create(
            father_name= father_name,
            father_occupation= father_occupation,
            father_mobile= father_mobile,
            father_email= father_email,
            mother_name= mother_name,
            mother_occupation= mother_occupation,
            mother_mobile= mother_mobile,
            mother_email= mother_email,
            present_address= present_address,
            permanent_address= permanent_address
        )

        class_obj = None
        if student_class:
            try:
                class_obj = Class.objects.get(pk=student_class)
            except Class.DoesNotExist:
                pass

        student = Student.objects.create(
            first_name= first_name,
            last_name= last_name,
            student_id= student_id,
            gender= gender,
            date_of_birth= date_of_birth,
            student_class= class_obj,
            religion= religion,
            joining_date= joining_date,
            mobile_number = mobile_number,
            admission_number = admission_number,
            section = section,
            student_image = student_image,
            parent = parent
        )
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã thêm học sinh: {student.last_name} {student.first_name}")
        messages.success(request, "Đã thêm học sinh thành công!")
        return redirect('student_list')

    classes = Class.objects.all()
    return render(request,"students/add-student.html", {'classes': classes})


def _normalize_header(name):
    return str(name).strip().lower().replace(' ', '_').replace('-', '_') if name is not None else ''


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_gender(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    if text in ('male', 'nam'):
        return 'Male'
    if text in ('female', 'nữ', 'nu'):
        return 'Female'
    if text in ('others', 'other', 'khác', 'khac'):
        return 'Others'
    return ''


def _parse_score(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(',', '.')
    try:
        score = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if score < 0:
        return None
    return score.quantize(Decimal('0.01'))


def _resolve_class(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        try:
            return Class.objects.get(pk=int(text))
        except Class.DoesNotExist:
            pass
    lookup = Class.objects.filter(class_name__iexact=text)
    if lookup.exists():
        return lookup.first()
    lookup = Class.objects.filter(class_code__iexact=text)
    if lookup.exists():
        return lookup.first()
    return None


@admin_required
def import_students_excel(request):
    classes = Class.objects.all()
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Vui lòng chọn file Excel để nhập.')
            return redirect('import_students_excel')

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception:
            messages.error(request, 'File Excel không hợp lệ hoặc không thể đọc được.')
            return redirect('import_students_excel')

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            messages.error(request, 'File Excel phải có ít nhất 1 dòng dữ liệu.')
            return redirect('import_students_excel')

        header_row = [_normalize_header(cell) for cell in rows[0]]
        required_columns = [
            'student_id', 'last_name', 'first_name', 'gender', 'date_of_birth', 'class',
            'joining_date', 'mobile_number', 'admission_number', 'father_name', 'father_mobile',
            'father_email', 'mother_name', 'mother_mobile', 'mother_email', 'present_address',
            'permanent_address'
        ]
        missing = [col for col in required_columns if col not in header_row]
        if missing:
            messages.error(request, 'File Excel thiếu cột: ' + ', '.join(missing))
            return redirect('import_students_excel')
        
        # Kiểm tra thứ tự cột: last_name phải đứng trước first_name
        last_name_idx = header_row.index('last_name')
        first_name_idx = header_row.index('first_name')
        if last_name_idx > first_name_idx:
            messages.error(request, 'Thứ tự cột không đúng. Cột "last_name" (Họ) phải đứng trước cột "first_name" (Tên). '
                          'Thứ tự đúng: student_id, last_name, first_name, gender, date_of_birth, class, religion, joining_date, '
                          'mobile_number, admission_number, section, father_name, father_occupation, father_mobile, father_email, '
                          'mother_name, mother_occupation, mother_mobile, mother_email, present_address, permanent_address')
            return redirect('import_students_excel')

        header_index = {name: idx for idx, name in enumerate(header_row)}
        created = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            student_id = row[header_index['student_id']]
            first_name = row[header_index['first_name']]
            last_name = row[header_index['last_name']]
            gender = _normalize_gender(row[header_index['gender']])
            date_of_birth = _parse_date(row[header_index['date_of_birth']])
            class_value = row[header_index['class']]
            religion = row[header_index['religion']] if 'religion' in header_index else ''
            joining_date = _parse_date(row[header_index['joining_date']])
            mobile_number = row[header_index['mobile_number']]
            admission_number = row[header_index['admission_number']]
            section = row[header_index['section']] if 'section' in header_index else ''
            father_name = row[header_index['father_name']]
            father_occupation = row[header_index['father_occupation']] if 'father_occupation' in header_index else ''
            father_mobile = row[header_index['father_mobile']]
            father_email = row[header_index['father_email']]
            mother_name = row[header_index['mother_name']]
            mother_occupation = row[header_index['mother_occupation']] if 'mother_occupation' in header_index else ''
            mother_mobile = row[header_index['mother_mobile']]
            mother_email = row[header_index['mother_email']]
            present_address = row[header_index['present_address']]
            permanent_address = row[header_index['permanent_address']]

            missing_fields = []
            if not student_id:
                missing_fields.append('student_id')
            if not first_name:
                missing_fields.append('first_name')
            if not last_name:
                missing_fields.append('last_name')
            if not gender:
                missing_fields.append('gender')
            if not date_of_birth:
                missing_fields.append('date_of_birth')
            if not joining_date:
                missing_fields.append('joining_date')
            if not mobile_number:
                missing_fields.append('mobile_number')
            if not admission_number:
                missing_fields.append('admission_number')
            if not father_name:
                missing_fields.append('father_name')
            if not father_mobile:
                missing_fields.append('father_mobile')
            if not father_email:
                missing_fields.append('father_email')
            if not mother_name:
                missing_fields.append('mother_name')
            if not mother_mobile:
                missing_fields.append('mother_mobile')
            if not mother_email:
                missing_fields.append('mother_email')
            if not present_address:
                missing_fields.append('present_address')
            if not permanent_address:
                missing_fields.append('permanent_address')

            if missing_fields:
                skipped += 1
                errors.append(f"Dòng {row_number}: thiếu {', '.join(missing_fields)}")
                continue

            class_obj = _resolve_class(class_value)
            if class_value and class_obj is None:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lớp '{class_value}' không tồn tại")
                continue

            if Student.objects.filter(student_id=student_id).exists():
                skipped += 1
                errors.append(f"Dòng {row_number}: Mã học sinh '{student_id}' đã tồn tại")
                continue
            if Student.objects.filter(admission_number=admission_number).exists():
                skipped += 1
                errors.append(f"Dòng {row_number}: Số nhập học '{admission_number}' đã tồn tại")
                continue

            try:
                with transaction.atomic():
                    parent = Parent.objects.create(
                        father_name=str(father_name).strip(),
                        father_occupation=str(father_occupation).strip() if father_occupation else '',
                        father_mobile=str(father_mobile).strip(),
                        father_email=str(father_email).strip(),
                        mother_name=str(mother_name).strip(),
                        mother_occupation=str(mother_occupation).strip() if mother_occupation else '',
                        mother_mobile=str(mother_mobile).strip(),
                        mother_email=str(mother_email).strip(),
                        present_address=str(present_address).strip(),
                        permanent_address=str(permanent_address).strip(),
                    )
                    Student.objects.create(
                        first_name=str(first_name).strip(),
                        last_name=str(last_name).strip(),
                        student_id=str(student_id).strip(),
                        gender=gender,
                        date_of_birth=date_of_birth,
                        student_class=class_obj,
                        religion=str(religion).strip() if religion else '',
                        joining_date=joining_date,
                        mobile_number=str(mobile_number).strip(),
                        admission_number=str(admission_number).strip(),
                        section=str(section).strip() if section else '',
                        parent=parent,
                    )
                created += 1
            except IntegrityError as exc:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lỗi dữ liệu hoặc giá trị trùng lặp ({exc})")
            except Exception as exc:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lỗi khi tạo học sinh ({exc})")

        if created:
            from school.models import Notification
            Notification.objects.create(user=request.user, message=f"Đã nhập {created} học sinh từ file Excel")

        if created > 0:
            messages.success(request, f"Đã nhập thành công {created} học sinh.")
        if skipped > 0:
            messages.warning(request, f"Bỏ qua {skipped} dòng. Xem chi tiết mô tả lỗi bên dưới.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"Còn {len(errors) - 10} lỗi khác.")

        return redirect('student_list')

    return render(request, 'students/import-students.html', {'classes': classes})


@admin_required
def import_teachers_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Vui lòng chọn file Excel để nhập.')
            return redirect('import_teachers_excel')

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception:
            messages.error(request, 'File Excel không hợp lệ hoặc không thể đọc được.')
            return redirect('import_teachers_excel')

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            messages.error(request, 'File Excel phải có ít nhất 1 dòng dữ liệu.')
            return redirect('import_teachers_excel')

        header_row = [_normalize_header(cell) for cell in rows[0]]
        required_columns = [
            'teacher_id', 'last_name', 'first_name', 'username', 'email'
        ]
        missing = [col for col in required_columns if col not in header_row]
        if missing:
            messages.error(request, 'File Excel thiếu cột: ' + ', '.join(missing))
            return redirect('import_teachers_excel')
        
        # Kiểm tra thứ tự cột: last_name phải đứng trước first_name
        last_name_idx = header_row.index('last_name')
        first_name_idx = header_row.index('first_name')
        if last_name_idx > first_name_idx:
            messages.error(request, 'Thứ tự cột không đúng. Cột "last_name" (Họ) phải đứng trước cột "first_name" (Tên). '
                          'Thứ tự đúng: teacher_id, last_name, first_name, username, email, phone_number, address, specialization, qualification, joining_date')
            return redirect('import_teachers_excel')

        header_index = {name: idx for idx, name in enumerate(header_row)}
        created = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            teacher_id = row[header_index['teacher_id']]
            first_name = row[header_index['first_name']]
            last_name = row[header_index['last_name']]
            username = row[header_index['username']]
            email = row[header_index['email']]
            phone_number = row[header_index['phone_number']] if 'phone_number' in header_index else ''
            address = row[header_index['address']] if 'address' in header_index else ''
            specialization = row[header_index['specialization']] if 'specialization' in header_index else ''
            qualification = row[header_index['qualification']] if 'qualification' in header_index else ''
            joining_date = _parse_date(row[header_index['joining_date']]) if 'joining_date' in header_index else None

            missing_fields = []
            if not teacher_id:
                missing_fields.append('teacher_id')
            if not first_name:
                missing_fields.append('first_name')
            if not last_name:
                missing_fields.append('last_name')
            if not username:
                missing_fields.append('username')
            if not email:
                missing_fields.append('email')

            if missing_fields:
                skipped += 1
                errors.append(f"Dòng {row_number}: thiếu {', '.join(missing_fields)}")
                continue

            if Teacher.objects.filter(teacher_id=teacher_id).exists():
                skipped += 1
                errors.append(f"Dòng {row_number}: Mã giáo viên '{teacher_id}' đã tồn tại")
                continue
            
            from home_auth.models import CustomUser
            if CustomUser.objects.filter(username=username).exists():
                skipped += 1
                errors.append(f"Dòng {row_number}: Tên đăng nhập '{username}' đã tồn tại")
                continue
            if CustomUser.objects.filter(email=email).exists():
                skipped += 1
                errors.append(f"Dòng {row_number}: Email '{email}' đã tồn tại")
                continue

            try:
                with transaction.atomic():
                    user = CustomUser.objects.create_user(
                        username=str(username).strip(),
                        email=str(email).strip(),
                        first_name=str(first_name).strip(),
                        last_name=str(last_name).strip(),
                        password='123456',  # Mật khẩu mặc định
                        is_teacher=True
                    )
                    Teacher.objects.create(
                        teacher_id=str(teacher_id).strip(),
                        user=user,
                        phone_number=str(phone_number).strip() if phone_number else '',
                        address=str(address).strip() if address else '',
                        specialization=str(specialization).strip() if specialization else '',
                        qualification=str(qualification).strip() if qualification else '',
                        joining_date=joining_date,
                    )
                created += 1
            except IntegrityError as exc:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lỗi dữ liệu hoặc giá trị trùng lặp ({exc})")
            except Exception as exc:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lỗi khi tạo giáo viên ({exc})")

        if created:
            from school.models import Notification
            Notification.objects.create(user=request.user, message=f"Đã nhập {created} giáo viên từ file Excel")

        if created > 0:
            messages.success(request, f"Đã nhập thành công {created} giáo viên.")
        if skipped > 0:
            messages.warning(request, f"Bỏ qua {skipped} dòng. Xem chi tiết mô tả lỗi bên dưới.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"Còn {len(errors) - 10} lỗi khác.")

        return redirect('teacher_list')

    return render(request, 'students/import-teachers.html')


@admin_required
def export_teachers_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mẫu giáo viên'

    headers = ['teacher_id', 'last_name', 'first_name', 'username', 'email', 'phone_number', 'address', 'specialization', 'qualification', 'joining_date']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sample_row = ['GV001', 'Nguyễn', 'Văn A', 'nguyenvana', 'nguyenvana@example.com', '0123456789', 'Hà Nội', 'Toán học', 'Thạc sĩ', datetime.now().strftime('%d/%m/%Y')]
    for col_num, value in enumerate(sample_row, 1):
        ws.cell(row=2, column=col_num, value=value)

    download_dir = r'C:\Users\Admin\Downloads'
    os.makedirs(download_dir, exist_ok=True)
    filename = 'Mau_Giao_Vien.xlsx'
    file_path = os.path.join(download_dir, filename)
    wb.save(file_path)

    messages.success(request, f'Đã xuất file mẫu vào {file_path}')
    return redirect('import_teachers_excel')


@admin_or_teacher_required
def export_teachers_excel(request):
    # Lấy danh sách giáo viên
    teachers = Teacher.objects.select_related('user').order_by('teacher_id')
    
    # Lọc theo trạng thái nếu có
    status = request.GET.get('status')
    if status == 'active':
        teachers = teachers.filter(is_active=True)
    elif status == 'inactive':
        teachers = teachers.filter(is_active=False)

    # Khởi tạo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Danh sách giáo viên'

    headers = ['STT', 'Mã giáo viên', 'Họ', 'Tên', 'Tên đăng nhập', 'Email', 'Số điện thoại', 'Địa chỉ', 'Chuyên môn', 'Trình độ', 'Ngày vào làm', 'Trạng thái']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Ghi Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Ghi dữ liệu giáo viên
    for row_num, teacher in enumerate(teachers, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=teacher.teacher_id)
        ws.cell(row=row_num, column=3, value=teacher.user.last_name)
        ws.cell(row=row_num, column=4, value=teacher.user.first_name)
        ws.cell(row=row_num, column=5, value=teacher.user.username)
        ws.cell(row=row_num, column=6, value=teacher.user.email)
        ws.cell(row=row_num, column=7, value=teacher.phone_number or '')
        ws.cell(row=row_num, column=8, value=teacher.address or '')
        ws.cell(row=row_num, column=9, value=teacher.specialization or '')
        ws.cell(row=row_num, column=10, value=teacher.qualification or '')
        ws.cell(row=row_num, column=11, value=teacher.joining_date.strftime('%d/%m/%Y') if teacher.joining_date else '')
        ws.cell(row=row_num, column=12, value='Đang làm việc' if teacher.is_active else 'Nghỉ việc')

    # Phản hồi tải file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Danh_sach_Giao_Vien_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin_required
def import_admission_candidates(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Vui lòng chọn file Excel để nhập.')
            return redirect('import_admission_candidates')

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception:
            messages.error(request, 'File Excel không hợp lệ hoặc không thể đọc được.')
            return redirect('import_admission_candidates')

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            messages.error(request, 'File Excel phải có ít nhất 1 dòng dữ liệu.')
            return redirect('import_admission_candidates')

        header_row = [_normalize_header(cell) for cell in rows[0]]
        required_columns = ['exam_number', 'full_name', 'date_of_birth', 'previous_school']
        missing = [col for col in required_columns if col not in header_row]
        if missing:
            messages.error(request, 'File Excel thiếu cột: ' + ', '.join(missing))
            return redirect('import_admission_candidates')

        header_index = {name: idx for idx, name in enumerate(header_row)}
        created = 0
        updated = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            exam_number = str(row[header_index['exam_number']]).strip() if header_index['exam_number'] < len(row) and row[header_index['exam_number']] is not None else ''
            full_name = str(row[header_index['full_name']]).strip() if header_index['full_name'] < len(row) and row[header_index['full_name']] is not None else ''
            date_of_birth = _parse_date(row[header_index['date_of_birth']]) if header_index['date_of_birth'] < len(row) else None
            previous_school = str(row[header_index['previous_school']]).strip() if header_index['previous_school'] < len(row) and row[header_index['previous_school']] is not None else ''

            if not exam_number or not full_name or not date_of_birth or not previous_school:
                skipped += 1
                errors.append(f"Dòng {row_number}: Thiếu dữ liệu bắt buộc.")
                continue

            try:
                candidate, created_flag = AdmissionCandidate.objects.update_or_create(
                    exam_number=exam_number,
                    defaults={
                        'full_name': full_name,
                        'date_of_birth': date_of_birth,
                        'previous_school': previous_school,
                    }
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Dòng {row_number}: Lỗi khi lưu thí sinh ({exc}).")

        if created > 0 or updated > 0:
            messages.success(request, f"Đã xử lý {created} thí sinh mới và cập nhật {updated} thí sinh.")
        if skipped > 0:
            messages.warning(request, f"Bỏ qua {skipped} dòng không hợp lệ.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"Còn {len(errors) - 10} lỗi khác.")

        return redirect('import_admission_candidates')

    return render(request, 'students/import-admissions.html')


@admin_required
def export_admission_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mẫu tuyển sinh'

    headers = ['exam_number', 'full_name', 'date_of_birth', 'previous_school']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sample_row = ['TS001', 'Nguyễn Văn A', datetime.now().strftime('%d/%m/%Y'), 'THCS Nguyễn Trãi']
    for col_num, value in enumerate(sample_row, 1):
        ws.cell(row=2, column=col_num, value=value)

    download_dir = r'C:\Users\Admin\Downloads'
    os.makedirs(download_dir, exist_ok=True)
    filename = 'Mau_Tuyen_Sinh_Lop_10.xlsx'
    file_path = os.path.join(download_dir, filename)
    wb.save(file_path)

    messages.success(request, f'Đã xuất file mẫu vào {file_path}')
    return redirect('import_admission_candidates')


@admin_required
def export_admission_scores(request):
    candidates = AdmissionCandidate.objects.all().order_by('exam_number')
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Diem tuyen sinh'

    headers = [
        'exam_number', 'full_name', 'date_of_birth', 'previous_school',
        'math_score', 'literature_score', 'english_score', 'total_score'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_num, candidate in enumerate(candidates, start=2):
        math_score = candidate.math_score
        literature_score = candidate.literature_score
        english_score = candidate.english_score
        total_score = None
        if any(value is not None for value in [math_score, literature_score, english_score]):
            total_score = sum((value or Decimal('0')) for value in [math_score, literature_score, english_score])

        values = [
            candidate.exam_number,
            candidate.full_name,
            candidate.date_of_birth.strftime('%d/%m/%Y') if candidate.date_of_birth else '',
            candidate.previous_school,
            math_score,
            literature_score,
            english_score,
            total_score,
        ]
        for col_num, value in enumerate(values, 1):
            sheet.cell(row=row_num, column=col_num).value = value

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Danh_sach_diem_tuyen_sinh.xlsx"'
    wb.save(response)
    return response


@admin_required
def assign_admission_classes(request):
    class_names = ['10A1', '10A2', '10A3']
    class_objects = []
    for class_name in class_names:
        class_obj, created = Class.objects.get_or_create(
            class_name=class_name,
            defaults={
                'class_code': class_name,
                'grade_level': '10',
                'capacity': 40,
            }
        )
        if class_obj.capacity != 40:
            class_obj.capacity = 40
            class_obj.save()
        class_objects.append(class_obj)

    candidates = AdmissionCandidate.objects.filter(
        math_score__isnull=False,
        literature_score__isnull=False,
        english_score__isnull=False,
    )

    ranked_candidates = sorted(
        candidates,
        key=lambda c: (
            -(c.total_score or Decimal('0')),
            -(c.math_score or Decimal('0')),
            -(c.literature_score or Decimal('0')),
            -(c.english_score or Decimal('0')),
            c.exam_number,
        )
    )

    total_required = 40 * len(class_objects)
    cutoff_candidate = None
    assigned_summary = None

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'assign':
            if len(ranked_candidates) < total_required:
                messages.error(request, f"Cần ít nhất {total_required} thí sinh có đủ điểm để phân chia 3 lớp.")
            else:
                AdmissionCandidate.objects.update(assigned_class=None)
                assigned = []
                for index, candidate in enumerate(ranked_candidates[:total_required]):
                    assigned_class = class_objects[index // 40]
                    candidate.assigned_class = assigned_class
                    candidate.save()
                    assigned.append(candidate)

                cutoff_candidate = assigned[-1]
                messages.success(request, f"Phân chia xong {total_required} thí sinh vào 3 lớp. Điểm chuẩn: {cutoff_candidate.total_score} (SBD: {cutoff_candidate.exam_number}).")
                return redirect('assign_admission_classes')
        elif action == 'sync':
            if not AdmissionCandidate.objects.filter(assigned_class__isnull=False).exists():
                messages.error(request, "Chưa có thí sinh nào được phân chia lớp để đồng bộ.")
            else:
                created_count, skipped_count, errors = _sync_admission_candidates(request.user)
                if created_count > 0:
                    sync_message = f"Đã đồng bộ {created_count} học sinh vào hệ thống."
                    if skipped_count > 0:
                        sync_message += f" {skipped_count} học sinh đã tồn tại và được bỏ qua."
                    messages.success(request, sync_message)
                elif skipped_count > 0:
                    messages.info(request, f"Tất cả {skipped_count} học sinh đã đồng bộ trước đó.")
                else:
                    messages.warning(request, "Không có học sinh mới nào được đồng bộ.")

                for error in errors:
                    messages.error(request, error)
                return redirect('assign_admission_classes')

    results_by_class = {}
    for class_obj in class_objects:
        results_by_class[class_obj.class_name] = list(
            AdmissionCandidate.objects.filter(assigned_class=class_obj).order_by(
                '-math_score', '-literature_score', '-english_score', 'exam_number'
            )
        )

    # Prepare class data for template
    class_data = []
    for class_obj in class_objects:
        class_data.append({
            'class_obj': class_obj,
            'students': results_by_class[class_obj.class_name],
            'count': len(results_by_class[class_obj.class_name])
        })

    if len(ranked_candidates) >= total_required:
        cutoff_candidate = ranked_candidates[total_required - 1]

    assigned_total = sum(data['count'] for data in class_data)

    return render(request, 'students/admission-assign-classes.html', {
        'class_data': class_data,
        'cutoff_candidate': cutoff_candidate,
        'ranked_count': len(ranked_candidates),
        'total_required': total_required,
        'assigned_total': assigned_total,
    })


def _sync_admission_candidates(user):
    assigned_candidates = AdmissionCandidate.objects.filter(assigned_class__isnull=False).select_related('assigned_class')
    created_count = 0
    skipped_count = 0
    errors = []

    for candidate in assigned_candidates:
        try:
            with transaction.atomic():
                if Student.objects.filter(admission_number=candidate.exam_number).exists():
                    skipped_count += 1
                    continue

                parent = Parent.objects.create(
                    father_name="Chưa cập nhật",
                    father_occupation="",
                    father_mobile="",
                    father_email="",
                    mother_name="Chưa cập nhật",
                    mother_occupation="",
                    mother_mobile="",
                    mother_email="",
                    present_address="Chưa cập nhật",
                    permanent_address="Chưa cập nhật",
                )

                Student.objects.create(
                    first_name=candidate.full_name.split()[-1] if candidate.full_name.split() else "Chưa cập nhật",
                    last_name=" ".join(candidate.full_name.split()[:-1]) if len(candidate.full_name.split()) > 1 else "",
                    student_id=candidate.exam_number,
                    gender='Male',
                    date_of_birth=candidate.date_of_birth,
                    student_class=candidate.assigned_class,
                    religion="",
                    joining_date=timezone.now().date(),
                    mobile_number="",
                    admission_number=candidate.exam_number,
                    section=candidate.assigned_class.grade_level if candidate.assigned_class else "",
                    parent=parent,
                    is_active=True,
                )
                created_count += 1
        except Exception as e:
            errors.append(f"Lỗi đồng bộ thí sinh {candidate.exam_number}: {str(e)}")

    if created_count > 0:
        from school.models import Notification
        Notification.objects.create(
            user=user,
            message=f"Đã đồng bộ {created_count} học sinh từ danh sách tuyển sinh"
        )
    return created_count, skipped_count, errors


@admin_required
def admission_scores(request):
    exam_number = request.GET.get('exam_number', '').strip()
    candidate = None

    # Xử lý import điểm hàng loạt
    if request.method == 'POST' and request.POST.get('bulk_import'):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Vui lòng chọn file Excel để nhập điểm.')
            return redirect('admission_scores')

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception:
            messages.error(request, 'File Excel không hợp lệ hoặc không thể đọc được.')
            return redirect('admission_scores')

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            messages.error(request, 'File Excel phải có ít nhất 1 dòng dữ liệu.')
            return redirect('admission_scores')

        header_row = [_normalize_header(cell) for cell in rows[0]]
        required_columns = ['exam_number', 'math_score', 'literature_score', 'english_score']
        missing = [col for col in required_columns if col not in header_row]
        if missing:
            messages.error(request, 'File Excel thiếu cột: ' + ', '.join(missing))
            return redirect('admission_scores')

        header_index = {name: idx for idx, name in enumerate(header_row)}
        updated = 0
        skipped = 0
        not_found = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            exam_number_row = str(row[header_index['exam_number']]).strip() if header_index['exam_number'] < len(row) and row[header_index['exam_number']] is not None else ''
            math_score = _parse_score(row[header_index['math_score']]) if header_index['math_score'] < len(row) else None
            literature_score = _parse_score(row[header_index['literature_score']]) if header_index['literature_score'] < len(row) else None
            english_score = _parse_score(row[header_index['english_score']]) if header_index['english_score'] < len(row) else None

            if not exam_number_row:
                skipped += 1
                errors.append(f"Dòng {row_number}: Thiếu số báo danh.")
                continue

            candidate = AdmissionCandidate.objects.filter(exam_number=exam_number_row).first()
            if not candidate:
                not_found += 1
                errors.append(f"Dòng {row_number}: Không tìm thấy thí sinh với số báo danh {exam_number_row}.")
                continue

            # Chỉ cập nhật điểm nếu ô đó chưa có dữ liệu
            updated_fields = []
            if math_score is not None and candidate.math_score is None:
                candidate.math_score = math_score
                updated_fields.append('Toán')
            if literature_score is not None and candidate.literature_score is None:
                candidate.literature_score = literature_score
                updated_fields.append('Văn')
            if english_score is not None and candidate.english_score is None:
                candidate.english_score = english_score
                updated_fields.append('Anh')

            if updated_fields:
                candidate.save()
                updated += 1
            else:
                skipped += 1

        if updated > 0:
            messages.success(request, f"Đã cập nhật điểm cho {updated} thí sinh.")
        if skipped > 0:
            messages.warning(request, f"Bỏ qua {skipped} dòng (đã có điểm hoặc thiếu dữ liệu).")
        if not_found > 0:
            messages.warning(request, f"Không tìm thấy {not_found} thí sinh trong danh sách.")
        if errors:
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"Còn {len(errors) - 10} lỗi khác.")

        return redirect('admission_scores')

    if request.method == 'POST' and not request.POST.get('bulk_import'):
        return redirect('admission_scores')

    if exam_number:
        candidate = AdmissionCandidate.objects.filter(exam_number=exam_number).first()
        if exam_number and not candidate:
            messages.error(request, f"Không tìm thấy thí sinh với số báo danh {exam_number}.")

    total_score = None
    if candidate:
        score_values = [candidate.math_score, candidate.literature_score, candidate.english_score]
        if any(value is not None for value in score_values):
            total_score = sum((value or Decimal('0')) for value in score_values)

    return render(request, 'students/admission-scores.html', {
        'candidate': candidate,
        'exam_number': exam_number,
        'total_score': total_score,
    })


@admin_or_teacher_required
def student_list(request):
    # Lấy danh sách tất cả các Khối duy nhất để hiển thị trong Dropdown
    all_grades = Class.objects.values_list('grade_level', flat=True).distinct().order_by('grade_level')
    selected_grade = request.GET.get('grade') # Lấy khối được chọn từ giao diện

    if request.user.is_admin:
        all_classes = Class.objects.all().order_by('grade_level', 'class_name')
        student_list = Student.objects.select_related('parent', 'student_class').filter(is_active=True)
    else:
        all_classes = Class.objects.filter(class_teacher=request.user).order_by('grade_level', 'class_name')
        student_list = Student.objects.filter(student_class__in=all_classes, is_active=True).select_related('parent', 'student_class')
    
    # Lọc danh sách theo Khối nếu người dùng chọn
    if selected_grade:
        all_classes = all_classes.filter(grade_level=selected_grade)
        student_list = student_list.filter(student_class__grade_level=selected_grade)

    # Giữ nguyên logic lọc theo Lớp (class_id) hiện tại của bạn...
    class_id = request.GET.get('class')
    selected_class = None
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            student_list = student_list.filter(student_class=selected_class)
        except Class.DoesNotExist:
            pass

    context = {
        'student_list': student_list.order_by('first_name', 'last_name'),
        'all_classes': all_classes,
        'all_grades': all_grades,         # Truyền biến này ra HTML
        'selected_grade': selected_grade, # Truyền biến này để giữ trạng thái Dropdown
        'selected_class': selected_class,
        'unread_notification': request.user.notification_set.filter(is_read=False)
    }
    return render(request, "students/students.html", context)

@admin_or_teacher_required
def export_students_excel(request):
    grade_level = request.GET.get('grade') # Lấy tham số 'grade' từ URL
    
    # 1. Xác định danh sách các Lớp cần xuất
    if request.user.is_admin:
        classes_to_export = Class.objects.all()
    else:
        classes_to_export = Class.objects.filter(class_teacher=request.user)

    if grade_level:
        classes_to_export = classes_to_export.filter(grade_level=grade_level)
    
    classes_to_export = classes_to_export.order_by('class_name')

    # 2. Khởi tạo Workbook
    wb = Workbook()
    
    # Xóa sheet mặc định ban đầu để tạo sheet theo tên lớp
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = ['STT', 'Mã học sinh', 'Họ', 'Tên', 'Giới tính', 'Ngày sinh', 'Lớp', 'Số điện thoại', 'Trạng thái']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # 3. Duyệt qua từng Lớp để tạo Sheet riêng
    for cls in classes_to_export:
        ws = wb.create_sheet(title=f"Lớp {cls.class_name}")
        
        # Ghi Header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Lấy học sinh của lớp này
        students = Student.objects.filter(student_class=cls, is_active=True).order_by('first_name')
        
        # Ghi dữ liệu học sinh vào sheet hiện tại
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=student.student_id)
            ws.cell(row=row_num, column=3, value=student.last_name)
            ws.cell(row=row_num, column=4, value=student.first_name)
            ws.cell(row=row_num, column=5, value=student.get_gender_display())
            ws.cell(row=row_num, column=6, value=student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '')
            ws.cell(row=row_num, column=7, value=cls.class_name)
            ws.cell(row=row_num, column=8, value=student.mobile_number)
            ws.cell(row=row_num, column=9, value='Đang học')

    # 4. Phản hồi tải file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Danh_sach_Khoi_{grade_level if grade_level else 'Tat_ca'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@admin_required
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    parent = student.parent if hasattr(student, 'parent') else None
    if request.method == "POST":
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        student_id = request.POST.get('student_id')
        gender = request.POST.get('gender')
        date_of_birth = request.POST.get('date_of_birth')
        student_class = request.POST.get('student_class')
        religion = request.POST.get('religion')
        joining_date = request.POST.get('joining_date')
        mobile_number = request.POST.get('mobile_number')
        admission_number = request.POST.get('admission_number')
        section = request.POST.get('section')
        student_image = request.FILES.get('student_image')  if request.FILES.get('student_image') else student.student_image

        parent.father_name = request.POST.get('father_name')
        parent.father_occupation = request.POST.get('father_occupation')
        parent.father_mobile = request.POST.get('father_mobile')
        parent.father_email = request.POST.get('father_email')
        parent.mother_name = request.POST.get('mother_name')
        parent.mother_occupation = request.POST.get('mother_occupation')
        parent.mother_mobile = request.POST.get('mother_mobile')
        parent.mother_email = request.POST.get('mother_email')
        parent.present_address = request.POST.get('present_address')
        parent.permanent_address = request.POST.get('permanent_address')
        parent.save()


        class_obj = None
        if student_class:
            try:
                class_obj = Class.objects.get(pk=student_class)
            except Class.DoesNotExist:
                pass
        
        student.first_name= first_name
        student.last_name= last_name
        student.student_id= student_id
        student.gender= gender
        student.date_of_birth= date_of_birth
        student.student_class= class_obj
        student.religion= religion
        student.joining_date= joining_date
        student.mobile_number = mobile_number
        student.admission_number = admission_number
        student.section = section
        student.student_image = student_image
        student.save()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã cập nhật học sinh: {student.last_name} {student.first_name}")
        messages.success(request, "Đã cập nhật học sinh thành công!")
        return redirect("student_list")
    
    classes = Class.objects.all()
    return render(request, "students/edit-student.html",{'student':student, 'parent':parent, 'classes': classes} )


@login_required
def view_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.user.is_student:
        pass
    elif request.user.is_teacher:
        if student.student_class and student.student_class.class_teacher != request.user:
            messages.error(request, "Bạn không có quyền xem thông tin học sinh này!")
            return redirect('student_list')
    
    context = {
        'student': student
    }
    return render(request, "students/student-details.html", context)


@admin_required
def delete_student(request, pk):
    if request.method == "POST":
        student = get_object_or_404(Student, pk=pk)
        student_name = f"{student.last_name} {student.first_name}"
        student.delete()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã xóa học sinh: {student_name}")
        messages.success(request, "Đã xóa học sinh thành công!")
        return redirect ('student_list')
    return HttpResponseForbidden()

@admin_or_teacher_required
def class_list(request):
    # Admin xem tất cả lớp, giáo viên chỉ xem lớp của mình
    if request.user.is_admin:
        classes = Class.objects.annotate(student_count=Count('students')).all().order_by('class_name')
    else:
        classes = Class.objects.filter(class_teacher=request.user).annotate(student_count=Count('students')).order_by('class_name')
    
    context = {
        'classes': classes,
    }
    return render(request, "students/class-list.html", context)

@admin_required
def add_class(request):
    if request.method == "POST":
        class_name = request.POST.get('class_name')
        class_code = request.POST.get('class_code')
        grade_level = request.POST.get('grade_level')
        capacity = request.POST.get('capacity', 30)
        class_teacher_id = request.POST.get('class_teacher')
        
        if Class.objects.filter(class_code=class_code).exists():
            messages.error(request, "Mã lớp đã tồn tại!")
            return render(request, "students/add-class.html")
        
        class_teacher = None
        if class_teacher_id:
            # Lấy user từ Teacher model
            teacher_obj = Teacher.objects.filter(id=class_teacher_id).first()
            if teacher_obj:
                class_teacher = teacher_obj.user
        
        Class.objects.create(
            class_name=class_name,
            class_code=class_code,
            grade_level=grade_level,
            capacity=capacity,
            class_teacher=class_teacher
        )
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã thêm lớp: {class_name}")
        messages.success(request, "Đã thêm lớp học thành công!")
        return redirect('class_list')
    
    teachers = Teacher.objects.filter(is_active=True).select_related('user')
    context = {'teachers': teachers}
    return render(request, "students/add-class.html", context)

@admin_required
def edit_class(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)
    if request.method == "POST":
        class_obj.class_name = request.POST.get('class_name')
        class_obj.class_code = request.POST.get('class_code')
        class_obj.grade_level = request.POST.get('grade_level')
        class_obj.capacity = request.POST.get('capacity', 30)
        class_teacher_id = request.POST.get('class_teacher')
        
        if class_teacher_id:
            teacher_obj = Teacher.objects.filter(id=class_teacher_id).first()
            if teacher_obj:
                class_obj.class_teacher = teacher_obj.user
            else:
                class_obj.class_teacher = None
        else:
            class_obj.class_teacher = None
        
        class_obj.save()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã cập nhật lớp: {class_obj.class_name}")
        messages.success(request, "Đã cập nhật lớp học thành công!")
        return redirect('class_list')
    
    teachers = Teacher.objects.filter(is_active=True).select_related('user')
    context = {'class_obj': class_obj, 'teachers': teachers}
    return render(request, "students/edit-class.html", context)

@admin_required
def delete_class(request, pk):
    if request.method == "POST":
        class_obj = get_object_or_404(Class, pk=pk)
        class_name = class_obj.class_name
        class_obj.delete()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã xóa lớp: {class_name}")
        messages.success(request, "Đã xóa lớp học thành công!")
        return redirect('class_list')
    return HttpResponseForbidden()

@admin_or_teacher_required
def subject_list(request):
    if request.user.is_admin:
        subjects = Subject.objects.select_related('teacher').all().order_by('subject_name')
    else:
        subjects = Subject.objects.filter(teacher=request.user).select_related('teacher').order_by('subject_name')
    
    context = {'subjects': subjects}
    return render(request, "students/subject-list.html", context)

@admin_required
def add_subject(request):
    if request.method == "POST":
        subject_name = request.POST.get('subject_name')
        subject_code = request.POST.get('subject_code')
        description = request.POST.get('description', '')
        teacher_id = request.POST.get('teacher')
        
        if Subject.objects.filter(subject_code=subject_code).exists():
            messages.error(request, "Mã môn học đã tồn tại!")
            return render(request, "students/add-subject.html")
        
        teacher = None
        if teacher_id:
            from home_auth.models import CustomUser
            teacher_obj = Teacher.objects.filter(id=teacher_id).first()
            if teacher_obj:
                teacher = teacher_obj.user
            else:
                teacher = None
        else:
            teacher = None
        
        Subject.objects.create(
            subject_name=subject_name,
            subject_code=subject_code,
            description=description,
            teacher=teacher
        )
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã thêm môn học: {subject_name}")
        messages.success(request, "Đã thêm môn học thành công!")
        return redirect('subject_list')
    
    teachers = Teacher.objects.filter(is_active=True).select_related('user')
    context = {'teachers': teachers}
    return render(request, "students/add-subject.html", context)

@admin_required
def edit_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == "POST":
        subject.subject_name = request.POST.get('subject_name')
        subject.subject_code = request.POST.get('subject_code')
        subject.description = request.POST.get('description', '')
        teacher_id = request.POST.get('teacher')
        
        if teacher_id:
            from home_auth.models import CustomUser
            teacher_obj = Teacher.objects.filter(id=teacher_id).first()
            if teacher_obj:
                subject.teacher = teacher_obj.user
            else:
                subject.teacher = None
        else:
            subject.teacher = None
        
        subject.save()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã cập nhật môn học: {subject.subject_name}")
        messages.success(request, "Đã cập nhật môn học thành công!")
        return redirect('subject_list')
    
    teachers = Teacher.objects.filter(is_active=True).select_related('user')
    context = {'subject': subject, 'teachers': teachers}
    return render(request, "students/edit-subject.html", context)

@admin_required
def delete_subject(request, pk):
    if request.method == "POST":
        subject = get_object_or_404(Subject, pk=pk)
        subject_name = subject.subject_name
        subject.delete()
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã xóa môn học: {subject_name}")
        messages.success(request, "Đã xóa môn học thành công!")
        return redirect('subject_list')
    return HttpResponseForbidden()

@login_required
def grade_list(request):
    if request.user.is_admin:
        grades = Grade.objects.select_related('student', 'subject').filter(is_approved=True)
        all_classes = Class.objects.all()
        subjects = Subject.objects.all()
    elif request.user.is_teacher:
        # Giáo viên chỉ xem điểm của học sinh trong lớp mình chủ nhiệm VÀ môn học mình dạy
        teacher_classes = Class.objects.filter(class_teacher=request.user)
        teacher_subjects = Subject.objects.filter(teacher=request.user)
        grades = Grade.objects.filter(
            student__student_class__in=teacher_classes,
            subject__in=teacher_subjects,
            is_approved=True
        ).select_related('student', 'subject')
        all_classes = teacher_classes
        subjects = teacher_subjects
    else:
        # Học sinh chỉ xem điểm của mình (tạm thời cho phép xem tất cả, cần cải thiện)
        grades = Grade.objects.select_related('student', 'subject').filter(is_approved=True)
        all_classes = Class.objects.all()
        subjects = Subject.objects.all()
    
    class_id = request.GET.get('class')
    subject_id = request.GET.get('subject')
    exam_type = request.GET.get('exam_type')
    
    selected_class = None
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            if request.user.is_teacher and selected_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền xem điểm của lớp này!")
                selected_class = None
            else:
                grades = grades.filter(student__student_class=selected_class)
        except Class.DoesNotExist:
            selected_class = None
    
    if subject_id:
        grades = grades.filter(subject_id=subject_id)
    if exam_type:
        grades = grades.filter(exam_type=exam_type)
    
    grades = grades.order_by('student__first_name', 'student__last_name', '-exam_date')
    
    context = {
        'grades': grades,
        'classes': all_classes,
        'subjects': subjects,
        'selected_class': selected_class,
    }
    return render(request, "students/grade-list.html", context)

@teacher_required
def add_grade(request):
    if request.method == "POST":
        from decimal import Decimal, InvalidOperation
        
        student_id = request.POST.get('student')
        subject_id = request.POST.get('subject')
        exam_type = request.POST.get('exam_type')
        score_str = request.POST.get('score')
        max_score_str = request.POST.get('max_score', '100')
        exam_date = request.POST.get('exam_date')
        remarks = request.POST.get('remarks', '')
        
        try:
            score = Decimal(str(score_str)) if score_str else Decimal('0')
            max_score = Decimal(str(max_score_str)) if max_score_str else Decimal('100')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Điểm số không hợp lệ: {str(e)}")
            # Lấy lại danh sách lớp, học sinh và môn học theo quyền
            if request.user.is_admin:
                all_classes = Class.objects.all().order_by('class_name')
                students = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
                subjects = Subject.objects.all().order_by('subject_name')
            else:
                all_classes = Class.objects.filter(class_teacher=request.user).order_by('class_name')
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True).order_by('first_name', 'last_name')
                subjects = Subject.objects.filter(teacher=request.user).order_by('subject_name')
            context = {
                'students': students, 
                'subjects': subjects,
                'all_classes': all_classes,
                'selected_class': None
            }
            return render(request, "students/add-grade.html", context)
        
        student = get_object_or_404(Student, id=student_id)
        subject = get_object_or_404(Subject, id=subject_id)
        
        if not request.user.is_admin:
            if not student.student_class or student.student_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền thêm điểm cho học sinh này! Học sinh không thuộc lớp bạn chủ nhiệm.")
                # Lấy lại danh sách lớp, học sinh và môn học theo quyền
                all_classes = Class.objects.filter(class_teacher=request.user).order_by('class_name')
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True).order_by('first_name', 'last_name')
                subjects = Subject.objects.filter(teacher=request.user).order_by('subject_name')
                context = {
                    'students': students, 
                    'subjects': subjects,
                    'all_classes': all_classes,
                    'selected_class': None
                }
                return render(request, "students/add-grade.html", context)
            
            if subject.teacher != request.user:
                messages.error(request, "Bạn không có quyền thêm điểm cho môn học này! Bạn không dạy môn học này.")
                all_classes = Class.objects.filter(class_teacher=request.user).order_by('class_name')
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True).order_by('first_name', 'last_name')
                subjects = Subject.objects.filter(teacher=request.user).order_by('subject_name')
                context = {
                    'students': students, 
                    'subjects': subjects,
                    'all_classes': all_classes,
                    'selected_class': None
                }
                return render(request, "students/add-grade.html", context)
        
        is_approved = request.user.is_admin
        
        grade = Grade.objects.create(
            student=student,
            subject=subject,
            exam_type=exam_type,
            score=score,
            max_score=max_score,
            exam_date=exam_date,
            remarks=remarks,
            is_approved=is_approved,
            approved_by=request.user if is_approved else None,
            approved_at=timezone.now() if is_approved else None
        )
        from school.models import Notification
        if is_approved:
            Notification.objects.create(user=request.user, message=f"Đã thêm điểm cho {student.first_name} - {subject.subject_name}")
            messages.success(request, "Đã thêm điểm thành công!")
        else:
            Notification.objects.create(user=request.user, message=f"Đã gửi yêu cầu thêm điểm cho {student.first_name} - {subject.subject_name}, đang chờ duyệt")
            messages.success(request, "Đã gửi yêu cầu thêm điểm thành công! Điểm sẽ được hiển thị sau khi được admin duyệt.")
        return redirect('grade_list')
    
    if request.user.is_admin:
        all_classes = Class.objects.all().order_by('class_name')
    else:
        # Giáo viên chỉ thấy lớp mình chủ nhiệm
        all_classes = Class.objects.filter(class_teacher=request.user).order_by('class_name')
    
    selected_class = None
    class_id = request.GET.get('class')
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            if request.user.is_teacher and selected_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền chọn lớp này!")
                selected_class = None
        except Class.DoesNotExist:
            selected_class = None
    
    if request.user.is_admin:
        if selected_class:
            students = Student.objects.filter(student_class=selected_class, is_active=True).order_by('first_name', 'last_name')
        else:
            students = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
        subjects = Subject.objects.all().order_by('subject_name')
    else:
        teacher_classes = Class.objects.filter(class_teacher=request.user)
        if selected_class:
            if selected_class in teacher_classes:
                students = Student.objects.filter(student_class=selected_class, is_active=True).order_by('first_name', 'last_name')
            else:
                students = Student.objects.none()
        else:
            students = Student.objects.filter(student_class__in=teacher_classes, is_active=True).order_by('first_name', 'last_name')
        subjects = Subject.objects.filter(teacher=request.user).order_by('subject_name')
    
    context = {
        'students': students, 
        'subjects': subjects,
        'all_classes': all_classes,
        'selected_class': selected_class
    }
    return render(request, "students/add-grade.html", context)

@teacher_required
def edit_grade(request, pk):
    grade = get_object_or_404(Grade, pk=pk)
    
    if not request.user.is_admin:
        if not grade.student.student_class or grade.student.student_class.class_teacher != request.user:
            messages.error(request, "Bạn không có quyền sửa điểm này! Học sinh không thuộc lớp bạn chủ nhiệm.")
            return redirect('grade_list')
        if grade.subject.teacher != request.user:
            messages.error(request, "Bạn không có quyền sửa điểm này! Bạn không dạy môn học này.")
            return redirect('grade_list')
    
    if request.method == "POST":
        from decimal import Decimal, InvalidOperation
        
        new_student_id = request.POST.get('student')
        new_subject_id = request.POST.get('subject')
        exam_type = request.POST.get('exam_type')
        
        if not request.user.is_admin:
            new_student = get_object_or_404(Student, id=new_student_id)
            new_subject = get_object_or_404(Subject, id=new_subject_id)
            
            if not new_student.student_class or new_student.student_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền sửa điểm này! Học sinh không thuộc lớp bạn chủ nhiệm.")
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True)
                subjects = Subject.objects.filter(teacher=request.user)
                context = {'grade': grade, 'students': students, 'subjects': subjects}
                return render(request, "students/edit-grade.html", context)
            
            if new_subject.teacher != request.user:
                messages.error(request, "Bạn không có quyền sửa điểm này! Bạn không dạy môn học này.")
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True)
                subjects = Subject.objects.filter(teacher=request.user)
                context = {'grade': grade, 'students': students, 'subjects': subjects}
                return render(request, "students/edit-grade.html", context)
        
        grade.student_id = new_student_id
        grade.subject_id = new_subject_id
        grade.exam_type = exam_type
        
        try:
            score_str = request.POST.get('score')
            max_score_str = request.POST.get('max_score', '100')
            grade.score = Decimal(str(score_str)) if score_str else Decimal('0')
            grade.max_score = Decimal(str(max_score_str)) if max_score_str else Decimal('100')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Điểm số không hợp lệ: {str(e)}")
            if request.user.is_admin:
                students = Student.objects.filter(is_active=True)
                subjects = Subject.objects.all()
            else:
                teacher_classes = Class.objects.filter(class_teacher=request.user)
                students = Student.objects.filter(student_class__in=teacher_classes, is_active=True)
                subjects = Subject.objects.filter(teacher=request.user)
            context = {'grade': grade, 'students': students, 'subjects': subjects}
            return render(request, "students/edit-grade.html", context)
        
        grade.exam_date = request.POST.get('exam_date')
        grade.remarks = request.POST.get('remarks', '')
        
        if request.user.is_admin:
            grade.is_approved = True
            grade.approved_by = request.user
            grade.approved_at = timezone.now()
        else:
            grade.is_approved = False
            grade.approved_by = None
            grade.approved_at = None
        
        grade.save()
        from school.models import Notification
        if grade.is_approved:
            Notification.objects.create(user=request.user, message=f"Đã cập nhật điểm cho {grade.student.first_name} - {grade.subject.subject_name}")
            messages.success(request, "Đã cập nhật điểm thành công!")
        else:
            Notification.objects.create(user=request.user, message=f"Đã gửi yêu cầu cập nhật điểm cho {grade.student.first_name} - {grade.subject.subject_name}, đang chờ duyệt")
            messages.success(request, "Đã gửi yêu cầu cập nhật điểm thành công! Điểm sẽ được hiển thị sau khi được admin duyệt.")
        return redirect('grade_list')
    
    if request.user.is_admin:
        students = Student.objects.filter(is_active=True)
        subjects = Subject.objects.all()
    else:
        # Giáo viên chỉ thấy học sinh trong lớp mình chủ nhiệm và môn học mình dạy
        teacher_classes = Class.objects.filter(class_teacher=request.user)
        students = Student.objects.filter(student_class__in=teacher_classes, is_active=True)
        subjects = Subject.objects.filter(teacher=request.user)
    
    context = {'grade': grade, 'students': students, 'subjects': subjects}
    return render(request, "students/edit-grade.html", context)

@teacher_required
def delete_grade(request, pk):
    grade = get_object_or_404(Grade, pk=pk)
    
    if not request.user.is_admin:
        if not grade.student.student_class or grade.student.student_class.class_teacher != request.user:
            messages.error(request, "Bạn không có quyền xóa điểm này! Học sinh không thuộc lớp bạn chủ nhiệm.")
            return redirect('grade_list')
        if grade.subject.teacher != request.user:
            messages.error(request, "Bạn không có quyền xóa điểm này! Bạn không dạy môn học này.")
            return redirect('grade_list')
    
    if request.method == "POST":
        grade.delete()
        messages.success(request, "Đã xóa điểm thành công!")
        return redirect('grade_list')
    return HttpResponseForbidden()

@admin_required
def approve_grades(request):
    """
    """
    all_classes = Class.objects.all().order_by('class_name')
    all_subjects = Subject.objects.all().order_by('subject_name')
    
    exam_types_in_db = Grade.objects.filter(is_approved=False).values_list('exam_type', flat=True).distinct().order_by('exam_type')
    
    class_id = request.GET.get('class')
    subject_id = request.GET.get('subject')
    exam_type = request.GET.get('exam_type')
    
    pending_grades = Grade.objects.filter(is_approved=False).select_related('student', 'subject', 'student__student_class')
    
    selected_class = None
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            pending_grades = pending_grades.filter(student__student_class=selected_class)
        except Class.DoesNotExist:
            selected_class = None
    
    selected_subject = None
    if subject_id:
        try:
            selected_subject = Subject.objects.get(id=subject_id)
            pending_grades = pending_grades.filter(subject=selected_subject)
        except Subject.DoesNotExist:
            selected_subject = None
    
    selected_exam_type = exam_type if exam_type else None
    if exam_type:
        pending_grades = pending_grades.filter(exam_type=exam_type)
    
    pending_grades = pending_grades.order_by('created_at')
    
    context = {
        'pending_grades': pending_grades,
        'all_classes': all_classes,
        'all_subjects': all_subjects,
        'exam_types_in_db': exam_types_in_db,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'selected_exam_type': selected_exam_type,
    }
    return render(request, "students/approve-grades.html", context)

@admin_required
def approve_grade(request, pk):
    """
    """
    grade = get_object_or_404(Grade, pk=pk)
    
    if request.method == "POST":
        grade.is_approved = True
        grade.approved_by = request.user
        grade.approved_at = timezone.now()
        grade.save()
        
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã duyệt điểm cho {grade.student.first_name} - {grade.subject.subject_name}")
        messages.success(request, f"Đã duyệt điểm cho {grade.student.first_name} - {grade.subject.subject_name} thành công!")
        
        params = []
        if request.GET.get('class'):
            params.append(f"class={request.GET.get('class')}")
        if request.GET.get('subject'):
            params.append(f"subject={request.GET.get('subject')}")
        if request.GET.get('exam_type'):
            params.append(f"exam_type={request.GET.get('exam_type')}")
        
        if params:
            return redirect(f"{reverse('approve_grades')}?{'&'.join(params)}")
        return redirect('approve_grades')
    
    return HttpResponseForbidden()

@admin_required
def reject_grade(request, pk):
    """
    """
    grade = get_object_or_404(Grade, pk=pk)
    
    if request.method == "POST":
        student_name = f"{grade.student.last_name} {grade.student.first_name}"
        subject_name = grade.subject.subject_name
        grade.delete()
        
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã từ chối điểm cho {student_name} - {subject_name}")
        messages.success(request, f"Đã từ chối điểm cho {student_name} - {subject_name}")
        
        # Giữ lại tất cả filter khi redirect
        params = []
        if request.GET.get('class'):
            params.append(f"class={request.GET.get('class')}")
        if request.GET.get('subject'):
            params.append(f"subject={request.GET.get('subject')}")
        if request.GET.get('exam_type'):
            params.append(f"exam_type={request.GET.get('exam_type')}")
        
        if params:
            return redirect(f"{reverse('approve_grades')}?{'&'.join(params)}")
        return redirect('approve_grades')
    
    return HttpResponseForbidden()

@admin_required
def approve_all_grades(request):

    if request.method == "POST":
        class_id = request.GET.get('class')
        subject_id = request.GET.get('subject')
        exam_type = request.GET.get('exam_type')
        
        pending_grades = Grade.objects.filter(is_approved=False)
        
        selected_class = None
        if class_id:
            try:
                selected_class = Class.objects.get(id=class_id)
                pending_grades = pending_grades.filter(student__student_class=selected_class)
            except Class.DoesNotExist:
                pass
        
        selected_subject = None
        if subject_id:
            try:
                selected_subject = Subject.objects.get(id=subject_id)
                pending_grades = pending_grades.filter(subject=selected_subject)
            except Subject.DoesNotExist:
                pass
        
        if exam_type:
            pending_grades = pending_grades.filter(exam_type=exam_type)
        
        count = pending_grades.count()
        
        if count > 0:
            now = timezone.now()
            pending_grades.update(
                is_approved=True,
                approved_by=request.user,
                approved_at=now
            )
            
            from school.models import Notification
            filter_info = []
            if selected_class:
                filter_info.append(f"lớp {selected_class.class_name}")
            if selected_subject:
                filter_info.append(f"môn {selected_subject.subject_name}")
            if exam_type:
                exam_type_choices = {
                    'Quiz': 'Kiểm tra 15 phút',
                    'Midterm': 'Giữa kỳ',
                    'Final': 'Cuối kỳ',
                    'Assignment': 'Bài tập',
                }
                exam_type_display = exam_type_choices.get(exam_type, exam_type)
                filter_info.append(f"loại {exam_type_display}")
            
            filter_text = f" ({', '.join(filter_info)})" if filter_info else ""
            Notification.objects.create(user=request.user, message=f"Đã duyệt tất cả {count} điểm chờ duyệt{filter_text}")
            messages.success(request, f"Đã duyệt thành công {count} điểm{filter_text}!")
        else:
            messages.info(request, "Không có điểm nào đang chờ duyệt.")
        
        params = []
        if class_id:
            params.append(f"class={class_id}")
        if subject_id:
            params.append(f"subject={subject_id}")
        if exam_type:
            params.append(f"exam_type={exam_type}")
        
        if params:
            return redirect(f"{reverse('approve_grades')}?{'&'.join(params)}")
        return redirect('approve_grades')
    
    return HttpResponseForbidden()

@login_required
def student_grades(request, pk):
    student = get_object_or_404(Student, pk=pk)
    grades = Grade.objects.filter(student=student, is_approved=True).select_related('subject').order_by('subject__subject_name', '-exam_date')
    
    subjects = Subject.objects.all().order_by('subject_name')
    subject_stats = []
    for subject in subjects:
        subject_grades = grades.filter(subject=subject)
        if subject_grades.exists():
            avg_score = subject_grades.aggregate(Avg('score'))['score__avg']
            subject_stats.append({
                'subject': subject,
                'average': round(avg_score, 2),
                'count': subject_grades.count()
            })
    
    context = {
        'student': student,
        'grades': grades,
        'subject_stats': subject_stats,
    }
    return render(request, "students/student-grades.html", context)

@login_required
def attendance_list(request):
    if request.user.is_admin:
        attendances = Attendance.objects.select_related('student').all()
        students = Student.objects.filter(is_active=True)
        all_classes = Class.objects.all()
    elif request.user.is_teacher:
        classes = Class.objects.filter(class_teacher=request.user)
        students = Student.objects.filter(student_class__in=classes, is_active=True)
        attendances = Attendance.objects.filter(student__in=students).select_related('student')
        all_classes = classes
    else:
        attendances = Attendance.objects.select_related('student').all()
        students = Student.objects.filter(is_active=True)
        all_classes = Class.objects.all()
    
    class_id = request.GET.get('class')
    student_id = request.GET.get('student')
    date = request.GET.get('date')
    status = request.GET.get('status')
    
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            if request.user.is_teacher and selected_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền xem điểm danh của lớp này!")
                selected_class = None
            else:
                students = students.filter(student_class=selected_class)
                attendances = attendances.filter(student__student_class=selected_class)
        except Class.DoesNotExist:
            selected_class = None
    else:
        selected_class = None
    
    if student_id:
        attendances = attendances.filter(student_id=student_id)
    if date:
        attendances = attendances.filter(date=date)
    if status:
        attendances = attendances.filter(status=status)
    
    attendances = attendances.order_by('student__first_name', 'student__last_name', '-date')
    
    context = {
        'attendances': attendances,
        'students': students,
        'classes': all_classes,
        'selected_class': selected_class,
    }
    return render(request, "students/attendance-list.html", context)

@teacher_required
def add_attendance(request):
    if request.method == "POST":
        date_str = request.POST.get('date')
        student_ids = request.POST.getlist('students')
        statuses = request.POST.getlist('status')
        remarks_list = request.POST.getlist('remarks')
        
        try:
            from datetime import datetime
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Ngày không hợp lệ!")
            return redirect('add_attendance')
        
        if attendance_date.isoweekday() == 7:
            messages.error(request, "Không thể điểm danh vào chủ nhật!")
            return redirect('add_attendance')
        
        for i, student_id in enumerate(student_ids):
            student = get_object_or_404(Student, id=student_id)
            status = statuses[i] if i < len(statuses) else 'Present'
            remarks = remarks_list[i] if i < len(remarks_list) else ''
            
            Attendance.objects.update_or_create(
                student=student,
                date=attendance_date,
                defaults={'status': status, 'remarks': remarks}
            )
        
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã điểm danh ngày {attendance_date}")
        messages.success(request, "Đã điểm danh thành công!")
        return redirect('attendance_list')
    
    selected_date = request.GET.get('date', timezone.now().date().isoformat())
    class_id = request.GET.get('class')
    
    if request.user.is_admin:
        classes = Class.objects.all()
    else:
        classes = Class.objects.filter(class_teacher=request.user)
    
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            if not request.user.is_admin and selected_class.class_teacher != request.user:
                messages.error(request, "Bạn không có quyền điểm danh lớp này!")
                selected_class = None
                students = Student.objects.none()
            else:
                students = Student.objects.filter(student_class=selected_class, is_active=True)
        except Class.DoesNotExist:
            selected_class = None
            students = Student.objects.none()
    else:
        selected_class = None
        if request.user.is_admin:
            students = Student.objects.filter(is_active=True)
        else:
            students = Student.objects.filter(student_class__in=classes, is_active=True)
    
    existing_attendance = {}
    if selected_date:
        for att in Attendance.objects.filter(date=selected_date, student__in=students):
            existing_attendance[att.student.id] = {'status': att.status, 'remarks': att.remarks}
    
    context = {
        'students': students,
        'classes': classes,
        'selected_date': selected_date,
        'selected_class': selected_class,
        'existing_attendance': existing_attendance,
    }
    return render(request, "students/add-attendance.html", context)

@teacher_required
def edit_attendance(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    
    if not request.user.is_admin:
        if not attendance.student.student_class or attendance.student.student_class.class_teacher != request.user:
            messages.error(request, "Bạn không có quyền sửa điểm danh này!")
            return redirect('attendance_list')
    attendance = get_object_or_404(Attendance, pk=pk)
    if request.method == "POST":
        attendance.date = request.POST.get('date')
        attendance.status = request.POST.get('status')
        attendance.remarks = request.POST.get('remarks', '')
        attendance.save()
        messages.success(request, "Đã cập nhật điểm danh thành công!")
        return redirect('attendance_list')
    
    context = {'attendance': attendance}
    return render(request, "students/edit-attendance.html", context)

@teacher_required
def delete_attendance(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    
    if not request.user.is_admin:
        if not attendance.student.student_class or attendance.student.student_class.class_teacher != request.user:
            messages.error(request, "Bạn không có quyền xóa điểm danh này!")
            return redirect('attendance_list')
    
    if request.method == "POST":
        attendance.delete()
        messages.success(request, "Đã xóa điểm danh thành công!")
        return redirect('attendance_list')
    return HttpResponseForbidden()

@login_required
def student_attendance(request, pk):
    student = get_object_or_404(Student, pk=pk)
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    
    total_days = attendances.count()
    present_count = attendances.filter(status='Present').count()
    absent_count = attendances.filter(status='Absent').count()
    late_count = attendances.filter(status='Late').count()
    excused_count = attendances.filter(status='Excused').count()
    
    attendance_rate = (present_count / total_days * 100) if total_days > 0 else 0
    
    context = {
        'student': student,
        'attendances': attendances,
        'total_days': total_days,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'attendance_rate': round(attendance_rate, 2),
    }
    return render(request, "students/student-attendance.html", context)

@admin_or_teacher_required
def reports_dashboard(request):
    # Student statistics
    total_students = Student.objects.filter(is_active=True).count()
    total_classes = Class.objects.count()
    total_subjects = Subject.objects.count()
    
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(date__gte=thirty_days_ago)
    attendance_rate = 0
    if recent_attendance.exists():
        present_count = recent_attendance.filter(status='Present').count()
        attendance_rate = (present_count / recent_attendance.count()) * 100
    
    recent_grades = Grade.objects.filter(exam_date__gte=thirty_days_ago)
    avg_score = recent_grades.aggregate(Avg('score'))['score__avg'] or 0
    
    class_stats = []
    for class_obj in Class.objects.all():
        student_count = class_obj.students.filter(is_active=True).count()
        
        students_in_class = class_obj.students.filter(is_active=True)
        average_grade = None
        if students_in_class.exists():
            # Lấy tất cả điểm của học sinh trong lớp
            grades_in_class = Grade.objects.filter(student__in=students_in_class)
            if grades_in_class.exists():
                avg_result = grades_in_class.aggregate(Avg('score'))['score__avg']
                if avg_result:
                    average_grade = float(avg_result)
        
        class_stats.append({
            'class': class_obj,
            'student_count': student_count,
            'capacity': class_obj.capacity,
            'fill_rate': (student_count / class_obj.capacity * 100) if class_obj.capacity > 0 else 0,
            'average_grade': average_grade
        })
    
    context = {
        'total_students': total_students,
        'total_classes': total_classes,
        'total_subjects': total_subjects,
        'attendance_rate': round(attendance_rate, 2),
        'avg_score': round(avg_score, 2),
        'class_stats': class_stats,
    }
    return render(request, "students/reports-dashboard.html", context)


@admin_required
def teacher_list(request):
    """
    Hiển thị danh sách tất cả giáo viên
    """
    teachers = Teacher.objects.select_related('user').order_by('teacher_id')
    
    # Lọc theo trạng thái
    status = request.GET.get('status')
    if status == 'active':
        teachers = teachers.filter(is_active=True)
    elif status == 'inactive':
        teachers = teachers.filter(is_active=False)
    
    # Tìm kiếm
    search_query = request.GET.get('search')
    if search_query:
        teachers = teachers.filter(
            Q(teacher_id__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(specialization__icontains=search_query)
        )
    
    context = {
        'teachers': teachers,
        'search_query': search_query,
        'status_filter': status,
        'unread_notification': request.user.notification_set.filter(is_read=False)
    }
    return render(request, 'students/teacher-list.html', context)

@admin_required
def add_teacher(request):
    from home_auth.models import CustomUser
    
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        phone_number = request.POST.get('phone_number', '')
        specialization = request.POST.get('specialization', '')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, "Tên đăng nhập đã tồn tại!")
            return render(request, "students/add-teacher.html")
        
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, "Email đã tồn tại!")
            return render(request, "students/add-teacher.html")
        
        try:
            user_account = CustomUser.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=password,
            )
            user_account.is_teacher = True
            user_account.is_authorized = True 
            user_account.save()
            

            teacher_id = f"GV{user_account.id:04d}"
            counter = 1
            while Teacher.objects.filter(teacher_id=teacher_id).exists():
                teacher_id = f"GV{user_account.id:04d}-{counter}"
                counter += 1
            
            teacher = Teacher.objects.create(
                teacher_id=teacher_id,
                user=user_account,
                phone_number=phone_number,
                address='',
                specialization=specialization,
                qualification='',
                joining_date=user_account.date_joined.date() if user_account.date_joined else None,
                is_active=True
            )
            
            from school.models import Notification
            Notification.objects.create(user=request.user, message=f"Đã thêm giáo viên: {teacher.get_full_name()}")
            messages.success(request, "Đã thêm giáo viên thành công!")
            return redirect('teacher_list')
        except Exception as e:
            messages.error(request, f"Có lỗi xảy ra: {str(e)}")
            return render(request, "students/add-teacher.html")
    
    return render(request, "students/add-teacher.html")

@admin_required
def edit_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    user = teacher.user
    
    if request.method == "POST":
        new_username = request.POST.get('username')
        new_email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone_number = request.POST.get('phone_number', '')
        specialization = request.POST.get('specialization', '')
        
        if user.username != new_username and user.__class__.objects.filter(username=new_username).exclude(pk=user.pk).exists():
            messages.error(request, "Tên đăng nhập đã tồn tại!")
            context = {'teacher': teacher}
            return render(request, "students/edit-teacher.html", context)
        
        if user.email != new_email and user.__class__.objects.filter(email=new_email).exclude(pk=user.pk).exists():
            messages.error(request, "Email đã tồn tại!")
            context = {'teacher': teacher}
            return render(request, "students/edit-teacher.html", context)
        
        user.username = new_username
        user.email = new_email
        user.first_name = first_name
        user.last_name = last_name
        user.save()
        
        teacher.phone_number = phone_number
        teacher.specialization = specialization
        teacher.save()
        
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
        
        user.save()
        
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã cập nhật thông tin giáo viên: {teacher.get_full_name()}")
        messages.success(request, "Đã cập nhật thông tin giáo viên thành công!")
        return redirect('teacher_list')
    
    context = {'teacher': teacher}
    return render(request, "students/edit-teacher.html", context)

@admin_required
def delete_teacher(request, pk):
    if request.method == "POST":
        teacher = get_object_or_404(Teacher, pk=pk)
        teacher_name = teacher.get_full_name()
        user = teacher.user
        
        teacher.delete()
        
        user.is_teacher = False
        user.save()
        
        from school.models import Notification
        Notification.objects.create(user=request.user, message=f"Đã xóa giáo viên: {teacher_name}")
        messages.success(request, "Đã xóa giáo viên thành công!")
        return redirect('teacher_list')
    return HttpResponseForbidden()


@admin_or_teacher_required
def student_list(request):
    # 1. Lấy danh sách tất cả các Khối duy nhất để hiển thị trong Dropdown
    all_grades = Class.objects.values_list('grade_level', flat=True).distinct().order_by('grade_level')
    
    # 2. Lấy giá trị khối và lớp được chọn từ request
    selected_grade = request.GET.get('grade')
    selected_class_id = request.GET.get('class')
    selected_class = None
    
    # 3. Logic lọc học sinh và lớp
    if request.user.is_admin:
        classes_query = Class.objects.all()
        student_query = Student.objects.filter(is_active=True)
    else:
        classes_query = Class.objects.filter(class_teacher=request.user)
        student_query = Student.objects.filter(student_class__in=classes_query, is_active=True)

    # 4. Áp dụng bộ lọc theo Khối nếu người dùng chọn
    if selected_grade:
        classes_query = classes_query.filter(grade_level=selected_grade)
        student_query = student_query.filter(student_class__grade_level=selected_grade)

    all_classes = classes_query.order_by('class_name')

    # 5. Áp dụng bộ lọc theo Lớp nếu người dùng chọn
    if selected_class_id:
        try:
            selected_class = all_classes.get(id=selected_class_id)
            student_query = student_query.filter(student_class=selected_class)
        except Class.DoesNotExist:
            selected_class = None

    student_list = student_query.select_related('parent', 'student_class').order_by('first_name', 'last_name')

    # 5. Truyền thêm các biến vào context
    context = {
        'student_list': student_list,
        'all_classes': all_classes,
        'all_grades': all_grades,         # Để lặp trong vòng for g in all_grades
        'selected_grade': selected_grade, # Để giữ trạng thái 'selected' trong dropdown
        'selected_class': selected_class,
        'unread_notification': request.user.notification_set.filter(is_read=False)
    }
    return render(request, "students/students.html", context)


@student_required
def student_dashboard(request):
    """
    Hiển thị dashboard cho học sinh
    """
    student = request.user.student_profile
    
    # Lấy thông tin điểm số gần đây
    recent_grades = Grade.objects.filter(student=student).select_related('subject').order_by('-created_at')[:5]
    
    # Lấy thông tin điểm danh gần đây
    recent_attendance = Attendance.objects.filter(student=student).order_by('-date')[:10]
    
    # Tính tỷ lệ điểm danh
    total_attendance = Attendance.objects.filter(student=student).count()
    present_count = Attendance.objects.filter(student=student, status='present').count()
    attendance_rate = (present_count / total_attendance * 100) if total_attendance > 0 else 0
    
    # Lấy thông tin lớp học
    student_class = student.student_class
    
    context = {
        'student': student,
        'recent_grades': recent_grades,
        'recent_attendance': recent_attendance,
        'attendance_rate': round(attendance_rate, 1),
        'student_class': student_class,
        'unread_notification': request.user.notification_set.filter(is_read=False)
    }
    return render(request, 'students/student-dashboard.html', context)


@login_required
def timetable_list(request):
    """Hiển thị thời khóa biểu theo quyền hạn"""
    
    # Admin: xem toàn bộ thời khóa biểu (giáo viên + học sinh)
    if request.user.is_admin:
        # Thời khóa biểu giáo viên
        teacher_schedules = Schedule.objects.all().select_related('teacher', 'class_obj', 'subject').order_by('teacher', 'day_of_week', 'period')
        
        # Group by teacher
        teacher_timetable = {}
        for schedule in teacher_schedules:
            teacher_name = f"{schedule.teacher.first_name} {schedule.teacher.last_name}"
            if teacher_name not in teacher_timetable:
                teacher_timetable[teacher_name] = {}
            key = f"{schedule.day_of_week}_{schedule.period}"
            teacher_timetable[teacher_name][key] = {
                'class': schedule.class_obj.class_name,
                'subject': schedule.subject.subject_name,
                'room': schedule.room
            }
        
        # Thời khóa biểu lớp
        class_schedules_list = Schedule.objects.all().select_related('teacher', 'class_obj', 'subject').order_by('class_obj__class_name', 'day_of_week', 'period')
        
        # Group by class
        class_timetable = {}
        for schedule in class_schedules_list:
            cls = schedule.class_obj.class_name
            if cls not in class_timetable:
                class_timetable[cls] = {}
            key = f"{schedule.day_of_week}_{schedule.period}"
            class_timetable[cls][key] = {
                'subject': schedule.subject.subject_name,
                'teacher': f"{schedule.teacher.first_name} {schedule.teacher.last_name}" if schedule.teacher else None,
                'room': schedule.room
            }
        
        context = {
            'is_admin': True,
            'teacher_timetable': teacher_timetable,
            'class_timetable': class_timetable,
            'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
            'periods': ['1', '2', '3', '4', '5']
        }
        return render(request, 'students/timetable-list.html', context)
    
    # Giáo viên: xem riêng thời khóa biểu của mình
    elif request.user.is_teacher:
        # Filter Schedule theo teacher user (vì Schedule.teacher là FK tới User)
        schedules = Schedule.objects.filter(teacher=request.user).select_related('class_obj', 'subject').order_by('day_of_week', 'period')
        
        timetable = {}
        for schedule in schedules:
            key = f"{schedule.day_of_week}_{schedule.period}"
            timetable[key] = {
                'class': schedule.class_obj.class_name,
                'subject': schedule.subject.subject_name,
                'room': schedule.room
            }
        
        context = {
            'is_teacher': True,
            'teacher_name': f"{request.user.first_name} {request.user.last_name}",
            'timetable': timetable,
            'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
            'periods': ['1', '2', '3', '4', '5']
        }
        return render(request, 'students/timetable-list.html', context)
    
    # Học sinh: xem riêng thời khóa biểu của lớp
    elif request.user.is_student:
        try:
            student = Student.objects.get(user=request.user)
            schedules = Schedule.objects.filter(class_obj=student.student_class).select_related('teacher', 'subject').order_by('day_of_week', 'period')
            
            timetable = {}
            for schedule in schedules:
                key = f"{schedule.day_of_week}_{schedule.period}"
                timetable[key] = {
                    'subject': schedule.subject.subject_name,
                    'teacher': f"{schedule.teacher.first_name} {schedule.teacher.last_name}" if schedule.teacher else None,
                    'room': schedule.room
                }
            
            context = {
                'is_student': True,
                'class_name': student.student_class.class_name if student.student_class else 'N/A',
                'timetable': timetable,
                'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
                'periods': ['1', '2', '3', '4', '5']
            }
            return render(request, 'students/timetable-list.html', context)
        except Student.DoesNotExist:
            messages.error(request, 'Không tìm thấy thông tin học sinh')
            return redirect('dashboard')
    
    messages.error(request, 'Bạn không có quyền truy cập')
    return redirect('dashboard')